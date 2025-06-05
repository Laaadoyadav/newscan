import requests
import pandas as pd
import time
from datetime import datetime
from typing import List
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class CVEDetails:
    def __init__(self, cve_id, description, cvss_score, severity, published_date, references):
        self.cve_id = cve_id
        self.description = description
        self.cvss_score = cvss_score
        self.severity = severity
        self.published_date = published_date
        self.references = references

class NVDApiClient:
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.base_url = "https://services.nvd.nist.gov/rest/json/cves/2.0"
        self.headers = {
            "X-Api-Key": api_key,
            "User-Agent": "Mozilla/5.0 (compatible; VulnerabilityScanner/1.0)"
        }
        self.request_delay = 6

    def get_cve_details(self, keyword: str, max_results: int = 5) -> List[CVEDetails]:
        try:
            params = {
                "keywordSearch": keyword,
                "resultsPerPage": max_results
            }

            time.sleep(self.request_delay)

            response = requests.get(self.base_url, headers=self.headers, params=params, timeout=30)

            if response.status_code == 403:
                print(f"API rate limit exceeded or authentication error: {response.text}")
                return []

            response.raise_for_status()
            data = response.json()

            cve_list = []
            for vuln in data.get("vulnerabilities", []):
                cve = vuln.get("cve", {})

                cvss_score = 0.0
                severity = "UNKNOWN"
                metrics = cve.get("metrics", {})
                if metrics.get("cvssMetricV31"):
                    cvss_data = metrics["cvssMetricV31"][0].get("cvssData", {})
                    cvss_score = float(cvss_data.get("baseScore", 0.0))
                    severity = cvss_data.get("baseSeverity", "UNKNOWN")
                elif metrics.get("cvssMetricV30"):
                    cvss_data = metrics["cvssMetricV30"][0].get("cvssData", {})
                    cvss_score = float(cvss_data.get("baseScore", 0.0))
                    severity = cvss_data.get("baseSeverity", "UNKNOWN")
                elif metrics.get("cvssMetricV2"):
                    cvss_data = metrics["cvssMetricV2"][0].get("cvssData", {})
                    cvss_score = float(cvss_data.get("baseScore", 0.0))
                    severity = "HIGH" if cvss_score >= 7 else "MEDIUM" if cvss_score >= 4 else "LOW"

                description = "No description available"
                for desc in cve.get("descriptions", []):
                    if desc.get("lang") == "en":
                        description = desc.get("value", "No description available")
                        break

                published_date = cve.get("published", "Unknown")
                references = [ref.get("url") for ref in cve.get("references", [])]

                cve_details = CVEDetails(
                    cve_id=cve.get("id", "Unknown"),
                    description=description,
                    cvss_score=cvss_score,
                    severity=severity,
                    published_date=published_date,
                    references=references
                )
                cve_list.append(cve_details)

            return cve_list

        except Exception as e:
            print(f"Error fetching CVE details: {e}")
            return []

# 🧠 Move EPSS lookup **outside** the class
def get_epss_details(cve_id):
    api_url = f"https://api.first.org/data/v1/epss?cve={cve_id}"
    try:
        response = requests.get(api_url, timeout=10)
        response.raise_for_status()
        data = response.json()
        item = data.get('data', [{}])[0]
        return {
            'cve': item.get('cve', cve_id),
            'epss': item.get('epss', 'N/A'),
            'percentile': item.get('percentile', 'N/A'),
            'date': item.get('date', 'N/A')
        }
    except Exception:
        return {
            'cve': cve_id,
            'epss': 'Error',
            'percentile': 'Error',
            'date': 'Error'
        }

# === MAIN SCRIPT ===

input_df = pd.read_excel("cve.xlsx")
keywords = input_df['keyword'].dropna().tolist()

api_key = "b0c2943a-8702-4560-b2b3-08ec65345d7b"  # Replace with your real key
client = NVDApiClient(api_key)
kev_url = "https://www.cisa.gov/sites/default/files/csv/known_exploited_vulnerabilities.csv"
kev_file = "known_exploited_vulnerabilities.csv"

response = requests.get(kev_url,timeout=10)
with open(kev_file, "wb") as f:
    f.write(response.content)

kev_df = pd.read_csv(kev_file)
kev_df['cveID'] = kev_df['cveID'].str.lower()
final_results = []

for keyword in keywords:
    cves = client.get_cve_details(keyword)
    print(f"🔍 Searching CVEs for keyword: {keyword}")
    for cve in cves:
        epss = get_epss_details(cve.cve_id)
        row = {
            'CVE ID': cve.cve_id,
            'Description': cve.description,
            'CVSS Score': cve.cvss_score,
            'Severity': cve.severity,
            'Published Date': cve.published_date,
            'References': ", ".join(cve.references),
            'EPSS Score': epss['epss'],
            'Percentile': epss['percentile'],
            'EPSS Date': epss['date']
        }
        final_results.append(row)

df = pd.DataFrame(final_results)

df['CVE ID'] = df['CVE ID'].str.lower()
df['In_KEV'] = df['CVE ID'].isin(kev_df['cveID']).map({True: 'Yes', False: 'No'})

# === Export to Excel ===
df.to_excel("output.xlsx", index=False)
# print("✅ Saved to output.xlsx")
wb = load_workbook("output.xlsx")
ws = wb.active

# === Define fills ===
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")       # CRITICAL / High EPSS
orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")    # HIGH / Medium EPSS
yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")    # MEDIUM
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  

# === Apply coloring ===
for row in ws.iter_rows(min_row=2, min_col=1, max_col=11):
    in_kev_cell = row[9]  # Column J is the 10th column (index 9)
    
    if in_kev_cell.value == "Yes":
        in_kev_cell.fill = red_fill
    elif in_kev_cell.value == "No":
        in_kev_cell.fill = green_fill
    
    # # Color by severity
    # if in_kev == 'Yes':
    #     if Cvss >=8:
    #         row[3].fill = red_fill
    # #     row[10].fill = red_fill
    # # else:
    # #     if epss == 'Error':
    # #         row[10].fill = red_fill
    # #     elif epss == 'N/A':
    # #         row[10].fill = yellow_fill
    # #     elif epss and float(epss) >= 0.5:
    # #         row[10].fill = red_fill
    # #     elif epss and float(epss) >= 0.2:
    # #         row[10].fill = orange_fill
    # #     elif epss and float(epss) >= 0.1:
    # #         row[10].fill = yellow_fill
    # #     else:
    # #         row[10].fill = green_fill
# === Save workbook ===
# date = datetime.now()
# formatted = date.strftime("%Y-%m-%d_%H-%M-%S")
filename = f"output_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
wb.save(filename)
print("✅ Excel with coloring saved as 'output .xlsx'")