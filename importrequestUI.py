import requests
import pandas as pd
import time
from datetime import datetime
from typing import List
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import Tk, filedialog, Text, Button, Label, END

# --- CVE Class ---
class CVEDetails:
    def __init__(self, cve_id, description, cvss_score, severity, published_date, references):
        self.cve_id = cve_id
        self.description = description
        self.cvss_score = cvss_score
        self.severity = severity
        self.published_date = published_date
        self.references = references

# --- NVD API Client ---
class NVDApiClient:
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.base_url = "https://services.nvd.nist.gov/rest/json/cves/2.0"
        self.headers = {
            "X-Api-Key": api_key,
            "User-Agent": "Mozilla/5.0 (compatible; VulnerabilityScanner/1.0)"
        }
        self.request_delay = 6

    def get_cve_details(self, keyword: str, max_results: int = 5) -> List[CVEDetails]:
        try:
            params = {
                "keywordSearch": keyword,
                "resultsPerPage": max_results
            }

            time.sleep(self.request_delay)
            response = requests.get(self.base_url, headers=self.headers, params=params, timeout=30)

            if response.status_code == 403:
                return []

            response.raise_for_status()
            data = response.json()

            cve_list = []
            for vuln in data.get("vulnerabilities", []):
                cve = vuln.get("cve", {})

                cvss_score = 0.0
                severity = "UNKNOWN"
                metrics = cve.get("metrics", {})
                if metrics.get("cvssMetricV31"):
                    cvss_data = metrics["cvssMetricV31"][0].get("cvssData", {})
                    cvss_score = float(cvss_data.get("baseScore", 0.0))
                    severity = cvss_data.get("baseSeverity", "UNKNOWN")
                elif metrics.get("cvssMetricV30"):
                    cvss_data = metrics["cvssMetricV30"][0].get("cvssData", {})
                    cvss_score = float(cvss_data.get("baseScore", 0.0))
                    severity = cvss_data.get("baseSeverity", "UNKNOWN")
                elif metrics.get("cvssMetricV2"):
                    cvss_data = metrics["cvssMetricV2"][0].get("cvssData", {})
                    cvss_score = float(cvss_data.get("baseScore", 0.0))
                    severity = "HIGH" if cvss_score >= 7 else "MEDIUM" if cvss_score >= 4 else "LOW"

                description = "No description available"
                for desc in cve.get("descriptions", []):
                    if desc.get("lang") == "en":
                        description = desc.get("value", "No description available")
                        break

                published_date = cve.get("published", "Unknown")
                references = [ref.get("url") for ref in cve.get("references", [])]

                cve_details = CVEDetails(
                    cve_id=cve.get("id", "Unknown"),
                    description=description,
                    cvss_score=cvss_score,
                    severity=severity,
                    published_date=published_date,
                    references=references
                )
                cve_list.append(cve_details)

            return cve_list

        except Exception as e:
            return []

# --- EPSS Fetcher ---
def get_epss_details(cve_id):
    api_url = f"https://api.first.org/data/v1/epss?cve={cve_id}"
    try:
        response = requests.get(api_url, timeout=10)
        response.raise_for_status()
        data = response.json()
        item = data.get('data', [{}])[0]
        return {
            'cve': item.get('cve', cve_id),
            'epss': item.get('epss', 'N/A'),
            'percentile': item.get('percentile', 'N/A'),
            'date': item.get('date', 'N/A')
        }
    except Exception:
        return {
            'cve': cve_id,
            'epss': 'Error',
            'percentile': 'Error',
            'date': 'Error'
        }

# === GUI Setup ===
def browse_file():
    filepath = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if filepath:
        label_file.config(text=filepath)
        run_button.config(state="normal")

def run_script():
    file_path = label_file.cget("text")
    output_box.insert(END, f"📥 Loaded file: {file_path}\n")
    input_df = pd.read_excel(file_path)
    keywords = input_df['keyword'].dropna().tolist()

    api_key = "b0c2943a-8702-4560-b2b3-08ec65345d7b"
    client = NVDApiClient(api_key)
    kev_url = "https://www.cisa.gov/sites/default/files/csv/known_exploited_vulnerabilities.csv"
    kev_file = "known_exploited_vulnerabilities.csv"

    response = requests.get(kev_url, timeout=10)
    with open(kev_file, "wb") as f:
        f.write(response.content)
    kev_df = pd.read_csv(kev_file)
    kev_df['cveID'] = kev_df['cveID'].str.lower()

    final_results = []
    for keyword in keywords:
        cves = client.get_cve_details(keyword)
        output_box.insert(END, f"🔍 Searching for keyword: {keyword}\n")
        output_box.see(END)
        for cve in cves:
            epss = get_epss_details(cve.cve_id)
            row = {
                'CVE ID': cve.cve_id,
                'Description': cve.description,
                'CVSS Score': cve.cvss_score,
                'Severity': cve.severity,
                'Published Date': cve.published_date,
                'References': ", ".join(cve.references),
                'EPSS Score': epss['epss'],
                'Percentile': epss['percentile'],
                'EPSS Date': epss['date']
            }
            final_results.append(row)
            output_box.insert(END, f"✅ Processed {cve.cve_id}\n")
            output_box.see(END)

    df = pd.DataFrame(final_results)
    df['CVE ID'] = df['CVE ID'].str.lower()
    df['In_KEV'] = df['CVE ID'].isin(kev_df['cveID']).map({True: 'Yes', False: 'No'})

    filename = f"output_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    df.to_excel(filename, index=False)

    wb = load_workbook(filename)
    ws = wb.active

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=11):
        in_kev_cell = row[9]
        if in_kev_cell.value == "Yes":
            in_kev_cell.fill = red_fill
        elif in_kev_cell.value == "No":
            in_kev_cell.fill = green_fill

    wb.save(filename)
    output_box.insert(END, f"\n✅ Excel saved: {filename}\n")
    output_box.see(END)

# === Build UI ===
root = Tk()
root.title("CVE EPSS Enricher")
root.geometry("600x500")

label_intro = Label(root, text="🔐 Upload CVE Keyword Excel file", font=("Arial", 12))
label_intro.pack(pady=10)

label_file = Label(root, text="No file selected", bg="white", width=60)
label_file.pack()

browse_button = Button(root, text="📂 Browse", command=browse_file)
browse_button.pack(pady=5)

run_button = Button(root, text="🚀 Run Analysis", state="disabled", command=run_script)
run_button.pack(pady=10)

output_box = Text(root, height=20, width=70)
output_box.pack()

root.mainloop()
